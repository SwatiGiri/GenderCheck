#!/usr/bin/env python
# coding: utf-8

# # Importing dependencies

# In[2]:


from simplesam import Reader, Writer
from pprint import pprint
import simplesam
import pysam
import glob
import pandas as pd
from openpyxl import Workbook, load_workbook
import os


    # # Importing all files from the folder

    # In[19]:

def Gendet(Runfolder, Samplesubsht):
        #change folder path name and add /*bam at the end of the path to import all bam files
    # Ydrive_path = "/media/sf_Y_DRIVE/"
    Ydrive_path = "/var/snap/docker/common/var-lib-docker/volumes/Ydrivedata/_data/"
    # Runfolder = input("Enter the name of the run folder to perform Gender QC: ")
    completepath = Ydrive_path + Runfolder
    print("Looking for bam files in : " + completepath + "....")
    # subdir = [x[0] for x in os.walk(Ydrive_path + Runfolder + "/")]
    # for i in subdir:
    # #     if '.bam' in file:
    #         print(file)
    files_list = [f for f in glob.glob(Ydrive_path + Runfolder + "/**/*.bam", recursive=True)]
    # files_list = [f for f in glob.glob(Ydrive_path + Runfolder + "/**/*.bam")]


    # print(files)
    for f in files_list:
        print(f)
    # Run39/190430_M05705_0046_000000000-C94MC_1/Alignment_Imported_1/20190502_120651/*.bam"
    # files_list = glob.glob(folder_path)


    # # Sample Names

    # In[21]:


    samples=[]
    for i in files_list:
        # master,media,Drive,Run_folder,seq_folder,alignement,final,file=i.split('/')
        lastindexfile = i.rsplit("/",1)

        print(lastindexfile)
        file = lastindexfile[1]
    #     master,folder,subfolder,file=i.split('/')
        sample,extension=file.split('_')
        samples.append(sample)
    samples


    # In[22]:


    len(samples)


    # # Opening the bam file to parse

    # In[23]:


    # file_path='/media/sf_Windows10Share/22647_S17.bam'


    # In[24]:


    # Caclulating reads from ChrY in location, 2654896 start position to 2655782 (this information can be seen in NCBI: https://www.ncbi.nlm.nih.gov/gene?term=NM_003140) stop position and storing in list
    XYratio=[]
    Gender=[]
    Reads=[]
    for i in files_list:
        samfile = pysam.AlignmentFile(i, "r")
        Read=samfile.count(contig='chrY',start=2654896,stop=2655782)
        Reads.append(Read)
        if Read >= 10:
            Gender.append("Male")
        else:
            Gender.append("Female")
        # count the number of reads in chrY
    #     Y=samfile.count(contig = "chrY")
    #     print(Y)
    #     # count the number of reads in chrX
    #     X=samfile.count(contig = "chrX")
    # #     print(X)
    #     # Ratio of number of reads of X chr and number of reads of Y chromosome
        
    #     ratio=(X+1)/(Y+1)
    #     if ratio > 13:
    #         Gender.append("Female")
    #     else :
    #         Gender.append("Male")
    #     XYratio.append(ratio)


    # In[25]:


    # XYratio
    Reads


    # In[26]:


    Gender


    # In[27]:


    # creating a dictionay containing samples and its sequenced predicted gender
    # SampleGender=dict(zip(samples,zip(XYratio,Gender)))
    # SampleGender=dict(zip(samples,Gender))
    SampleGender=dict(zip(samples,zip(Reads,Gender)))
    # SampleGender


    # In[28]:


    # df=pd.DataFrame.from_dict(SampleGender,orient='index',columns=['XYratio','Gender'])
    df=pd.DataFrame.from_dict(SampleGender,orient='index',columns=['Reads','Gender'])
    df


    # #  Creating a folder to save Gender results

    # In[31]:


    # Change folder path name to the root project and add "gender" folder at the end
    try:
        os.stat(completepath + "/Gender")
    except:
        os.makedirs(completepath + "/Gender")
    # Saving csv file containing Gender determined by sequenicng
    df.to_csv(completepath + '/Gender/GenDet.csv')


    # In[32]:


    SampleGender=dict(zip(samples,Gender))


    # In[34]:


    # Importing xlsx file containing gender information from the submission sheet
    # Samplesubsht = input("Enter the name of the sample submission sheet : ")
    subshtpath = completepath + "/" + Samplesubsht
    wb= load_workbook(subshtpath)
    ws=wb.active
    a=ws["B15"].value
    print(a)


    # In[35]:


    # change row range based on the number of rows in sample submission sheet to get all the samples
    row_range=ws['B15':'B68']
    InitSamples=[]
    for row in row_range:
            for cell in row:
                    InitSamples.append(cell.value)
    InitSamples


    # In[36]:


    # chnage row range based on the number of rows in sample submission sheet
    row_range=ws['E15':'E68']
    InitGender=[]
    for row in row_range:
            for cell in row:
                    InitGender.append(cell.value)
    InitGender


    # In[37]:


    InitData=dict(zip(InitSamples,InitGender))
    print(InitData)


    # In[38]:


    # Finding difference and saving the results in text file
    def compare_dict(dict1, dict2):
        for x1 in dict1.keys():
            z = dict1.get(x1) == dict2.get(x1)
            if not z:
                print('key', x1)
                print('value A', dict1.get(x1), '\nvalue B', dict2.get(x1))
                print('-----\n')
                if dict2.get(x1) != None:
                    file=open(completepath + '/Gender/GenDif.txt','a')
                    file.write('Sample: '+str(x1)+ ' Known: '+str(dict1.get(x1))+' Sequenced: '+str(dict2.get(x1))+"\n")
                    file.close

    compare_dict(InitData,SampleGender)

# Gendet('NGSValidationAccuracy', 'NGSValidationAccuracyLIMSSampleSubmissionSheet.xlsx')
    # In[39]:


    # # Returns true if index is present
    # samfile.check_index()
    # # count the number of reads in a region
    # samfile.count()
    # # count the number of reads in chrY
    # Y=samfile.count(contig = "chrY")
    # # count the number of reads in chrX
    # X=samfile.count(contig = "chrX")


    # In[40]:


    # # Ratio of number of reads of X chr and number of reads of Y chromosome
    # ratio=X/Y
    # print(ratio)


    # In[41]:


    # iterate=samfile.fetch('chrY')
    # for read in iterate:
    #      print(read)

    # samfile.close()


    # In[ ]:


    # file_path='/media/sf_Run33/22767_S13.bam'
    # in_file=open(file_path,'r')
    # in_bam = Reader(in_file)


    # In[ ]:


    # parsed_file=pysam.AlignmentFile(file_path,'r')
    # Y=parsed_file.count(start=2654896,stop=2655782)
    # print(Y)


    # Access alignments using an iterator interface

    # In[ ]:


    # pprint(in_bam.header)


    # In[ ]:


    # for i in in_bam:
    #     print(i)


    # In[ ]:


    # x=next(in_bam)
    # type(x)


    # In[ ]:


    # x.qname


    # # In[ ]:


    # x.rname


    # # In[ ]:


    # x.pos


    # # In[ ]:


    # x.seq


    # # In[ ]:


    # x.qual


    # # In[ ]:


    # x.cigar


    # # In[ ]:


    # x.gapped


    # # In[ ]:


    # x.flag


    # # In[ ]:


    # x.mapped


    # # In[ ]:


    # x.duplicate


    # # In[ ]:


    # x.secondary


    # # In[ ]:


    # x.tags

